from collections.abc import AsyncGenerator
from typing import IO

from .page import Page
from .parser import Parser

from io import BytesIO
from openpyxl import load_workbook



class ExcelParser(Parser):
    """

    """

    async def parse(self, content: IO) -> AsyncGenerator[Page, None]:
        blob_data = content
        blob_stream = BytesIO(blob_data.read())
        workbook = load_workbook(blob_stream, data_only=True)
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            yield Page(page_num=sheet_name, offset=None, text=sheet)
 